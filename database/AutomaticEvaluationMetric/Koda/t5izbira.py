from transformers import AutoModelForMaskedLM, AutoTokenizer
import torch
import numpy as np

model_name = 'EMBEDDIA/sloberta'
model = AutoModelForMaskedLM.from_pretrained(model_name)
tokenizer = AutoTokenizer.from_pretrained(model_name)

def score(model, tokenizer, sentence):
    tensor_input = tokenizer.encode(sentence, return_tensors='pt')
    repeat_input = tensor_input.repeat(tensor_input.size(-1)-2, 1)
    mask = torch.ones(tensor_input.size(-1) - 1).diag(1)[:-2]
    masked_input = repeat_input.masked_fill(mask == 1, tokenizer.mask_token_id)
    labels = repeat_input.masked_fill( masked_input != tokenizer.mask_token_id, -100)
    with torch.inference_mode():
        loss = model(masked_input, labels=labels).loss
    return np.exp(loss.item())


from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('/media/Podatki/FRI/NLP/Projekt/nlp-course-parafrizerji/RezultatiParafraze/T5_Google_150.xlsx')

sheet = workbook['Sheet1']

seznam_stavkov = []
j = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(j)
    j+=1
    row = row[1:]
    if row[1] != None:
        row = list(filter(lambda item: item is not None, row))
        indeks = 1
        najmanj = score(sentence=row[1], model=model, tokenizer=tokenizer)
        for i in range(2,len(row)):
            per_score = score(sentence=row[i], model=model, tokenizer=tokenizer)
            if per_score < najmanj:
                najmanj = per_score
                indeks = i
        seznam_stavkov.append([row[0], row[indeks]])
    else:
        seznam_stavkov.append([row[0], ""])
    
workbook.close()

workbook_out = Workbook()

sheet = workbook_out.active

for row in seznam_stavkov:
    sheet.append(row)

workbook_out.save('T5_Google_150_izbran.xlsx')
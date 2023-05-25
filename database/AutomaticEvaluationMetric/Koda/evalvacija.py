import evaluate
sacrebleu = evaluate.load("sacrebleu")
meteor = evaluate.load('meteor')
    
from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('/media/Podatki/FRI/NLP/evalvacija/izbran_stavek/T5_Google_200_izbran.xlsx')

sheet = workbook['Sheet']

seznam_stavkov = []
j = 0
for row in sheet.iter_rows(min_row=1, values_only=True):
    print(j)
    j+=1
    #row = row[1:]
    if row[1] == None:
        seznam_stavkov.append([0, 0])
    else:
        predictions = [row[1]]
        references = [row[0]]
        results_bleu = sacrebleu.compute(predictions=predictions, references=references)
        results_meteor = meteor.compute(predictions=predictions, references=references)
        seznam_stavkov.append([round(results_bleu["score"]/100, 3), round(results_meteor['meteor'], 3)])
    
workbook.close()

workbook_out = Workbook()

sheet = workbook_out.active

for row in seznam_stavkov:
    sheet.append(row)

workbook_out.save('T5_Google_200_eval.xlsx')
